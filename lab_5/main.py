import logging
import os
import sys
from lab_5.solver import Solver
import xlrd
import math
import openpyxl
from copy import deepcopy

from PyQt5.uic import loadUiType

import sys
from PyQt5.QtCore import pyqtSlot, pyqtSignal, QTimer
from PyQt5.QtWidgets import QApplication,QTableWidgetItem, QDialog, QFileDialog, QMessageBox

app = QApplication(sys.argv)
app.setApplicationName('LABA 5')
form_class, base_class = loadUiType('lab_5/main_window_2.ui')

data = {'col1':['1','2','3','4'],
        'col2':['1','2','1','3'],
        'col3':['1','1','2','1']}


class MainWindow_2(QDialog, form_class):
    # signals:
    input_changed = pyqtSignal('QString')
    output_changed = pyqtSignal('QString')


    def __init__(self, parent = None):
        super(MainWindow_2, self).__init__(parent)



        # setting up ui
        self.setupUi(self)

    
        self.setWindowTitle('Військово-промисловий комплекс')
        self.input_path = self.line_input.text()
        if len(self.input_path) == 0:
            self.input_path = os_path
        self.output_path = './output.xlsx'

        #set tablewidget
        self.tablewidget.verticalHeader().hide()
        self.tablewidget_first.verticalHeader().hide()
        self.tablewidget_second.verticalHeader().hide()
        self.tablewidget.horizontalHeader().hide()
        self.tablewidget_first.horizontalHeader().hide()
        self.tablewidget_second.horizontalHeader().hide()
        #self.tablewidget.setRowCount(0)
        '''
        column_size = [60, 70, 100, 100,200, 60, 200, 80]
        for index, size in enumerate(column_size):
             self.tablewidget.setColumnWidth(index,size)
        '''
        return

    @pyqtSlot()
    def input_clicked(self):
        filename = QFileDialog.getOpenFileName(self, 'Відкрити данні', '.', 'Data file (*.xlsx)')[0]
        if filename == '':
            return
        if filename != self.input_path:
            self.input_path = filename
            self.input_changed.emit(filename)
        return

    @pyqtSlot('QString')
    def input_modified(self, value):
        if value != self.input_path:
            self.input_path = value
        return



    @pyqtSlot()
    def output_clicked(self):
        filename = QFileDialog.getSaveFileName(self, 'Save data file', '.', 'Spreadsheet (*.xlsx)')[0]
        if filename == '':
            return
        if filename != self.output_path:
            self.output_path = filename
            self.output_changed.emit(filename)
        return

    @pyqtSlot('QString')
    def output_modified(self, value):
        if value != self.output_path:
            self.output_path = value
        return

    
    @pyqtSlot()
    def exec_clicked(self):
        self.exec_button.setEnabled(False)
        try:
            prob = 0.5

            p_in = xlrd.open_workbook(self.input_path)
            p_out = openpyxl.Workbook()

            s = Solver(prob)
            s.load_data(p_in)

            tmp, output1 = s.solve_task1(p_out.create_sheet('Task 1'), self.tablewidget_first)
            tmp, output2 = s.solve_task2(p_out.create_sheet('Task 2'), tmp,self.tablewidget_second)
            tmp, output3 = s.solve_task3(p_out.create_sheet('Task 3'), tmp, self.tablewidget)

            p_out.save(self.output_path)

        except Exception as e:
            QMessageBox.warning(self,'Error!','Error happened during execution: ' + str(e))
        self.exec_button.setEnabled(True)
        return