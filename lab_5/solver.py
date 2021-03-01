import xlrd
import math
import openpyxl
from copy import deepcopy
import numpy as np
from PyQt5.QtWidgets import QTableWidgetItem


class Solver:
    def __init__(self, p):
        self.AlternativesNames = []
        self.Alternatives = []
        self.OutcomeNames = []
        self.Outcomes = []

        self.Alt_Expert_Prob = []
        self.Alternatives_Dependencies = []
        self.Outcomes_Dependencies = []
        self.Alternatives_Outcomes_Crossdependencies = []

        self.p = p

    def solve_task1(self, outsheet, table):
        for i in range(len(self.Alternatives)):
            summ = 0
            for j in range(len(self.Alternatives[i])):
                summ = summ + self.Alt_Expert_Prob[i][j]
            for j in range(len(self.Alternatives[i])):
                self.Alt_Expert_Prob[i][j] = self.Alt_Expert_Prob[i][j] / summ
        alt_prob = self.modify_probabilities(self.Alt_Expert_Prob, self.Alternatives_Dependencies)

        row = 0
        column = 0
        for i in range(len(self.Alternatives)):
            outsheet.cell(1, 3*i+1).value = self.AlternativesNames[i]
            for j in range(len(self.Alternatives[i])):
                outsheet.cell(2+j, 3*i+1).value = self.Alternatives[i][j] + ' :'
                outsheet.cell(2+j, 3*i+2).value = str(alt_prob[i][j])
                if 2 + j > row:
                    row = 2 + j
                if 3 * i + 2 > column:
                    column = 3 * i + 2

        result = np.empty((row, column), dtype='object')
        for i in range(len(self.Alternatives)):
            newitem = QTableWidgetItem(str(self.AlternativesNames[i]))
            table.setItem(0, 3*i,newitem)
            result[0, 3*i] = self.AlternativesNames[i]
            for j in range(len(self.Alternatives[i])):
                newitem = QTableWidgetItem(str(self.Alternatives[i][j] + ' :'))
                table.setItem(1+j, 3*i,newitem )
                newitem = QTableWidgetItem(str(alt_prob[i][j]))
                table.setItem(1+j, 3*i+1, newitem)
                result[1+j, 3*i] = self.Alternatives[i][j] + ' :'
                result[1+j, 3*i+1] = str(alt_prob[i][j])
        table.resizeColumnsToContents()
        table.resizeRowsToContents()
        return alt_prob, result

    def solve_task2(self, outsheet, alt_prob, table):
        outcome_prob = self.calculate_outcome_prob(alt_prob)
        outcomes_c = [None for k in range(len(self.Outcomes))]
        for i in range(len(self.Outcomes)):
            outcomes_c[i] = deepcopy(self.Outcomes[i])
            sort = sorted(zip(outcome_prob[i], outcomes_c[i]), reverse=True)
            outcomes_c[i] = [x for y, x in sort]
            outcome_prob[i] = [y for y, x in sort]

        row = 0
        column = 0
        for i in range(len(self.Outcomes)):
            outsheet.cell(1, 3*i+1).value = self.OutcomeNames[i]
            for j in range(len(self.Outcomes[i])):
                outsheet.cell(2+j, 3*i+1).value = outcomes_c[i][j] + ' :'
                outsheet.cell(2+j, 3*i+2).value = str(outcome_prob[i][j])
                if 2 + j > row:
                    row = 2 + j
                if 3 * i + 2 > column:
                    column = 3 * i + 2

        result = np.empty((row, column), dtype='object')
        for i in range(len(self.Outcomes)):
            newitem = QTableWidgetItem(str(self.OutcomeNames[i]))
            table.setItem(0, 3*i,newitem)
            result[0, 3*i] = self.OutcomeNames[i]
            for j in range(len(self.Outcomes[i])):
                newitem = QTableWidgetItem(str(outcomes_c[i][j] + ' :'))
                table.setItem(1+j, 3*i,newitem )
                newitem = QTableWidgetItem(str(outcome_prob[i][j]))
                table.setItem(1+j, 3*i+1, newitem)
                result[1+j, 3*i] = outcomes_c[i][j] + ' :'
                result[1+j, 3*i+1] = str(outcome_prob[i][j])
        table.resizeColumnsToContents()
        table.resizeRowsToContents()

        return outcome_prob, result

    def solve_task3(self, outsheet, outcome_prob, table):
        outcome_mod_prob = self.modify_probabilities(outcome_prob, self.Outcomes_Dependencies)
        outcomes_c = [None for k in range(len(self.Outcomes))]
        for i in range(len(self.Outcomes)):
            outcomes_c[i] = deepcopy(self.Outcomes[i])
            sort = sorted(zip(outcome_mod_prob[i], outcomes_c[i]), reverse=True)
            outcomes_c[i] = [x for y, x in sort]
            outcome_mod_prob[i] = [y for y, x in sort]

        row = 0
        column = 0
        for i in range(len(self.Outcomes)):
            outsheet.cell(1, 3 * i + 1).value = self.OutcomeNames[i]
            for j in range(len(self.Outcomes[i])):
                outsheet.cell(2 + j, 3 * i + 1).value = outcomes_c[i][j] + ' :'
                outsheet.cell(2 + j, 3 * i + 2).value = str(outcome_mod_prob[i][j])
            outsheet.cell(2, 3*i+1).font = openpyxl.styles.Font(bold=True)
            outsheet.cell(2, 3*i+2).font = openpyxl.styles.Font(bold=True)
            if 2 + j > row:
                row = 2 + j
            if 3 * i + 2 > column:
                column = 3 * i + 2

        result = np.empty((row, column), dtype='object')
        for i in range(len(self.Outcomes)):
            newitem = QTableWidgetItem(str(self.OutcomeNames[i]))
            table.setItem(0, 3*i,newitem)
            result[0, 3*i] = self.OutcomeNames[i]
            for j in range(len(self.Outcomes[i])):
                newitem = QTableWidgetItem(str(outcomes_c[i][j] + ' :'))
                table.setItem(1+j, 3*i,newitem )
                newitem = QTableWidgetItem(str(outcome_mod_prob[i][j]))
                table.setItem(1+j, 3*i+1, newitem)

                result[1+j, 3*i] = outcomes_c[i][j] + ' :'
                result[1+j, 3*i+1] = str(outcome_mod_prob[i][j])
        table.resizeColumnsToContents()
        table.resizeRowsToContents()

        
        return outcome_mod_prob, result

    def modify_probabilities(self, starting_prob, dep_matrix):
        cur = 0
        l = len(starting_prob)
        res = [None for k in range(2)]
        res[0] = [None for k in range(l)]
        res[1] = [None for k in range(l)]

        for i in range(l):
            res[0][i] = [None for k in range(len(starting_prob[i]))]
            res[1][i] = [None for k in range(len(starting_prob[i]))]

        for i in range(l):
            for j in range(len(starting_prob[i])):
                res[1][i][j] = starting_prob[i][j]

        for iteration in range(500):
            for i in range(l):
                for j in range(len(starting_prob[i])):
                    n = [0 for k in range(l)]
                    res[cur][i][j] = 0
                    while True:
                        # calculating of probabilities
                        cur_dep = [None for k in range(l - 1)]
                        for k in range(l):
                            if k > i:
                                cur_dep[k - 1] = dep_matrix[i][k][j][n[k]]
                            elif k < i:
                                cur_dep[k] = dep_matrix[i][k][j][n[k]]
                        p_a = self.calculate_p(cur_dep)

                        for k in range(l):
                            if k == i:
                                continue
                            p_a = p_a * res[1 - cur][k][n[k]]
                        res[cur][i][j] = res[cur][i][j] + p_a

                        # modification of index array
                        cn = l - 1
                        while cn >= 0:
                            n[cn] = n[cn] + 1
                            if n[cn] == len(starting_prob[cn]):
                                n[cn] = 0
                                cn = cn - 1
                            else:
                                break
                        if cn < 0:
                            break

                # norming the probabilities
                summ = 0
                for j in range(len(starting_prob[i])):
                    summ = summ + res[cur][i][j]
                for j in range(len(starting_prob[i])):
                    res[cur][i][j] = res[cur][i][j] / summ

            cur = 1 - cur
        return res[cur]

    def calculate_outcome_prob(self, alt_prob):
        res = [None for k in range(len(self.Outcomes))]
        for i in range(len(res)):
            res[i] = [0 for k in range(len(self.Outcomes[i]))]
        for i in range(len(self.Outcomes)):
            for j in range(len(self.Outcomes[i])):
                n = [None for k in range(len(self.Alternatives))]
                for k in range(len(self.Alternatives)):
                    n[k] = 0
                while True:
                    # calculating of probabilities
                    cur_dep = [None for k in range(len(self.Alternatives))]
                    for k in range(len(self.Alternatives)):
                        cur_dep[k] = self.Alternatives_Outcomes_Crossdependencies[k][i][n[k]][j]
                    p_a = self.calculate_p(cur_dep)
                    for k in range(len(self.Alternatives)):
                        if k == i:
                            continue
                        p_a = p_a * alt_prob[k][n[k]]
                    res[i][j] = res[i][j] + p_a

                    #modification of index array
                    cn = len(self.Outcomes) - 1
                    while cn >= 0:
                        n[cn] = n[cn] + 1
                        if n[cn] == len(self.Alternatives[cn]):
                            n[cn] = 0
                            cn = cn - 1
                        else:
                            break
                    if cn < 0:
                        break
            summ = 0
            for j in range(len(self.Outcomes[i])):
                summ = summ + res[i][j]
            for j in range(len(self.Outcomes[i])):
                res[i][j] = res[i][j] / summ
        return res

    def calculate_p(self, a_i):
        c = 1
        for i in range(len(a_i)):
            c = c * (2 / (1 - a_i[i]) - 1)
        c = 1 - 2 / (c + 1)

        rp = math.cos((math.acos(1 - 2 * self.p) + math.pi) / 3) + 0.5
        if rp <= 0.5:
            np = -1 * math.log(rp, 2)
        else:
            np = -1 / math.log(1 - rp, 2)
        if np >= 1:
            tc = 1 - 2 * math.pow((1 - c) / 2, np)
        else:
            tc = 2 * math.pow((1 + c) / 2, 1 / np) - 1
        return 3 * ((tc + 1) / 2) * ((tc + 1) / 2) - 2 * ((tc + 1) / 2) * ((tc + 1) / 2) * ((tc + 1) / 2)

    def load_data(self, file):
        sheet = file.sheet_by_index(0)

        # inputting alternatives
        alt_count, out_count = None, None
        for i in range(1000):
            if str(sheet.row_values(0)[i]) == '*':
                alt_count = i
                break
        max_row = 0

        for i in range(alt_count):
            self.AlternativesNames.append(sheet.row_values(0)[i])
            max_ind = None
            for j in range(1, 1000):
                if sheet.row_values(j+1)[i] == '*':
                    max_ind = j
                    break
            self.Alternatives.append([])
            self.Alt_Expert_Prob.append([])
            for j in range(max_ind):
                self.Alternatives[i].append(sheet.row_values(j+1)[i])
                self.Alt_Expert_Prob[i].append(None)
            if max_row < max_ind:
                max_row = max_ind

        # inputting outcomes
        base_row = max_row + 2
        for i in range(1000):
            if sheet.row_values(base_row)[i] == '*':
                out_count = i
                break
        max_row = 0
        for i in range(out_count):
            self.OutcomeNames.append(sheet.row_values(base_row)[i])
            max_ind = None
            for j in range(1000):
                if sheet.row_values(j + 1 + base_row)[i] == '*':
                    max_ind = j
                    break
            self.Outcomes.append([])
            for j in range(max_ind):
                self.Outcomes[i].append(sheet.row_values(j + 1 + base_row)[i])
            if max_row < max_ind:
                max_row = max_ind
        base_row = base_row + max_row + 3

        # inputting alternative probabilities
        max_row = 0
        for i in range(alt_count):
            for j in range(len(self.Alt_Expert_Prob[i])):
                self.Alt_Expert_Prob[i][j] = float(sheet.row_values(base_row + j)[i])
        base_row = 1

        # inputting alt-alt dependency matrix
        sheet = file.sheet_by_index(1)
        self.Alternatives_Dependencies = [[None for h in range(alt_count)] for k in range(alt_count)]
        for i in range(alt_count - 1):
            base_column = 1
            for j in range(1, alt_count):
                if j <= i:
                    base_column = base_column + len(self.Alternatives[j])
                    continue
                li = len(self.Alternatives[i])
                lj = len(self.Alternatives[j])
                self.Alternatives_Dependencies[i][j] = [[None for h in range(lj)] for k in range(li)]
                self.Alternatives_Dependencies[j][i] = [[None for h in range(li)] for k in range(lj)]
                for i1 in range(li):
                    for j1 in range(lj):
                        if sheet.row_values(base_row + i1)[base_column + j1] == '*':
                            self.Alternatives_Dependencies[i][j][i1][j1] = 0
                            self.Alternatives_Dependencies[j][i][j1][i1] = 0
                        else:
                            self.Alternatives_Dependencies[i][j][i1][j1] = float(
                                sheet.row_values(base_row + i1)[base_column + j1])
                            self.Alternatives_Dependencies[j][i][j1][i1] = float(
                                sheet.row_values(base_row + i1)[base_column + j1])
                base_column = base_column + len(self.Alternatives[j])
            base_row = base_row + len(self.Alternatives[i])
        base_row = base_row + 2

        # inputting alt_out dependency matrix
        self.Alternatives_Outcomes_Crossdependencies = [[None for h in range(out_count)] for k in range(alt_count)]
        for i in range(alt_count):
            base_column = 1
            for j in range(out_count):
                li = len(self.Alternatives[i])
                lj = len(self.Outcomes[j])
                self.Alternatives_Outcomes_Crossdependencies[i][j] = [[None for h in range(lj)] for k in range(li)]
                for i1 in range(li):
                    for j1 in range(lj):
                        if sheet.row_values(base_row + i1)[base_column + j1] == '*':
                            self.Alternatives_Outcomes_Crossdependencies[i][j][i1][j1] = 0
                        else:
                            self.Alternatives_Outcomes_Crossdependencies[i][j][i1][j1] = float(
                                sheet.row_values(base_row + i1)[base_column + j1])
                base_column = base_column + len(self.Outcomes[j])
            base_row = base_row + len(self.Alternatives[i])
        base_row = base_row + 2

        # inputting out-out dependency matrix
        self.Outcomes_Dependencies = [[None for h in range(out_count)] for k in range(out_count)]
        for i in range(out_count):
            base_column = 1
            for j in range(1, out_count):
                if j <= i:
                    base_column = base_column + len(self.Outcomes[j])
                    continue
                li = len(self.Outcomes[i])
                lj = len(self.Outcomes[j])
                self.Outcomes_Dependencies[i][j] = [[None for h in range(lj)] for k in range(li)]
                self.Outcomes_Dependencies[j][i] = [[None for h in range(li)] for k in range(lj)]
                for i1 in range(li):
                    for j1 in range(lj):
                        if sheet.row_values(base_row + i1)[base_column + j1] == '*':
                            self.Outcomes_Dependencies[i][j][i1][j1] = 0
                            self.Outcomes_Dependencies[j][i][j1][i1] = 0
                        else:
                            self.Outcomes_Dependencies[i][j][i1][j1] = float(
                                sheet.row_values(base_row + i1)[base_column + j1])
                            self.Outcomes_Dependencies[j][i][j1][i1] = float(
                                sheet.row_values(base_row + i1)[base_column + j1])
                base_column = base_column + len(self.Outcomes[j])
            base_row = base_row + len(self.Outcomes[i])


